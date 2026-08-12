#!/usr/bin/env python3
"""
Cria leads.xlsx com aba Leads (template) e aba Dashboard.
Execute uma vez antes de começar a campanha.
Você pode importar sua lista de leads por qualquer uma dessas formas:
  - Copiar/colar diretamente na aba Leads (a partir da linha 2)
  - Importar CSV via Excel (Dados → De Texto/CSV)
"""

import openpyxl
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference
from datetime import date

OUTPUT = "leads.xlsx"

# ── Cores ──────────────────────────────────────────────────────────────────────
COR_AZUL_ESC = "1F4E79"
COR_AZUL_MED = "2E75B6"
COR_VERDE = "375623"
COR_VERDE_CL = "E2EFDA"
COR_LARANJA = "C55A11"
COR_LARANJA_C = "FCE4D6"
COR_CINZA = "F2F2F2"
COR_VERMELHO = "C00000"
COR_VERM_CL = "FFE7E6"
COR_AMARELO = "FFD966"
COR_AMARELO_C = "FFFACD"
COR_BRANCO = "FFFFFF"

# Cores por status (para validação visual)
COR_STATUS = {
    "pendente": ("595959", "F2F2F2"),  # (fonte, fundo)
    "abordado": ("1F4E79", "D9E8F5"),
    "ignorado": ("7F6000", "FFF2CC"),
    "recusou": ("C00000", "FFE7E6"),
    "interessado": ("375623", "E2EFDA"),
}


def borda_fina():
    s = Side(style="thin", color="BFBFBF")
    return Border(left=s, right=s, top=s, bottom=s)


def cabecalho_cell(ws, row, col, valor, cor_fundo=COR_AZUL_ESC, cor_fonte=COR_BRANCO, negrito=True):
    c = ws.cell(row, col, valor)
    c.fill = PatternFill("solid", fgColor=cor_fundo)
    c.font = Font(color=cor_fonte, bold=negrito, size=10)
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    c.border = borda_fina()
    return c


# ── Aba Leads ──────────────────────────────────────────────────────────────────
def criar_aba_leads(wb: openpyxl.Workbook):
    ws = wb.active
    ws.title = "Leads"
    ws.sheet_view.showGridLines = True

    # Colunas: (título, largura, formato)
    colunas = [
        ("Nome", 28, "@"),
        ("Telefone", 20, "@"),
        ("Cidade", 18, "@"),
        ("Status", 14, "@"),
        ("1º Contato", 14, "DD/MM/YYYY"),
        ("Último Contato", 16, "DD/MM/YYYY"),
        ("Próxima Tentativa", 18, "DD/MM/YYYY"),
        ("Tentativas", 13, "0"),
        ("Última Resposta", 40, "@"),
        ("Hora Resposta", 18, "DD/MM/YYYY HH:MM"),
        ("Observações", 35, "@"),
    ]

    ws.row_dimensions[1].height = 32

    for idx, (titulo, larg, fmt) in enumerate(colunas, 1):
        cabecalho_cell(ws, 1, idx, titulo)
        ws.column_dimensions[get_column_letter(idx)].width = larg

    # Congelar linha de cabeçalho
    ws.freeze_panes = "A2"

    # A planilha e criada VAZIA de proposito: contatos de exemplo ja causaram
    # envios de teste indevidos. Preencha a partir da linha 2 com leads reais.

    # Validação de dados para coluna Status (D)
    from openpyxl.worksheet.datavalidation import DataValidation
    dv = DataValidation(
        type="list",
        formula1='"pendente,abordado,ignorado,recusou,interessado"',
        showDropDown=False,
        error="Status inválido. Use: pendente, abordado, ignorado, recusou, interessado",
        errorTitle="Status inválido",
        showErrorMessage=True,
    )
    dv.sqref = "D2:D10000"
    ws.add_data_validation(dv)

    # Filtro automático
    ws.auto_filter.ref = f"A1:{get_column_letter(len(colunas))}1"

    # Instrução no cabeçalho da planilha
    ws.cell(1, 12).value = "← Preencha a partir daqui com seus leads. Telefone: 55DDNÚMERO (ex: 5521999990001)"
    ws.cell(1, 12).font = Font(italic=True, color="7F7F7F", size=9)
    ws.column_dimensions["L"].width = 60


# ── Aba Dashboard ──────────────────────────────────────────────────────────────
def criar_aba_dashboard(wb: openpyxl.Workbook):
    ws = wb.create_sheet("Dashboard")
    ws.sheet_view.showGridLines = False

    # Título
    ws.merge_cells("A1:D1")
    c = ws["A1"]
    c.value = "📊 Dashboard — Campanha WhatsApp"
    c.font = Font(bold=True, size=16, color=COR_AZUL_ESC)
    c.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 36

    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 18

    def linha_metrica(row, label, valor, cor_fundo=COR_CINZA, cor_fonte="000000", negrito=False):
        ws.row_dimensions[row].height = 22
        a = ws.cell(row, 1, label)
        a.fill = PatternFill("solid", fgColor=cor_fundo)
        a.font = Font(bold=True, size=10, color="444444")
        a.alignment = Alignment(vertical="center", indent=1)
        a.border = borda_fina()

        b = ws.cell(row, 2, valor)
        b.fill = PatternFill("solid", fgColor=cor_fundo)
        b.font = Font(bold=negrito, size=12, color=cor_fonte)
        b.alignment = Alignment(horizontal="center", vertical="center")
        b.border = borda_fina()

    # Seção: Geral
    ws.merge_cells("A3:D3")
    t = ws["A3"]
    t.value = "VISÃO GERAL"
    t.font = Font(bold=True, size=10, color=COR_BRANCO)
    t.fill = PatternFill("solid", fgColor=COR_AZUL_MED)
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[3].height = 20

    linha_metrica(4, "Total Leads", 0, COR_CINZA, "000000")
    linha_metrica(5, "Total Abordados", 0, "D9E8F5", COR_AZUL_ESC, True)
    linha_metrica(6, "Abordados Hoje", 0, "EBF3FB", COR_AZUL_MED)
    linha_metrica(7, "Taxa Resposta (%)", 0.0, COR_CINZA, "000000")
    linha_metrica(8, "Última Atualização", "-", COR_CINZA, "595959")

    # Seção: Funil
    ws.merge_cells("A10:D10")
    t2 = ws["A10"]
    t2.value = "FUNIL DE PROSPECÇÃO"
    t2.font = Font(bold=True, size=10, color=COR_BRANCO)
    t2.fill = PatternFill("solid", fgColor=COR_AZUL_MED)
    t2.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[10].height = 20

    funil = [
        ("Pendentes", 0, "F2F2F2", "595959"),
        ("Abordados", 0, "D9E8F5", COR_AZUL_ESC),
        ("Ignorados", 0, "FFF2CC", "7F6000"),
        ("Recusaram", 0, "FFE7E6", COR_VERMELHO),
        ("Interessados", 0, "E2EFDA", COR_VERDE),
    ]
    for i, (label, val, bg, fg) in enumerate(funil, 11):
        linha_metrica(i, label, val, bg, fg, True)
        ws.row_dimensions[i].height = 24

    # Legenda de status
    ws.merge_cells("A17:D17")
    t3 = ws["A17"]
    t3.value = "SIGNIFICADO DOS STATUS"
    t3.font = Font(bold=True, size=10, color=COR_BRANCO)
    t3.fill = PatternFill("solid", fgColor="595959")
    t3.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[17].height = 20

    legendas = [
        ("pendente", "Ainda não foi contactado"),
        ("abordado", "Mensagem enviada, aguardando resposta ou recontato"),
        ("ignorado", "Não respondeu nas 2 tentativas — encerrado automaticamente"),
        ("recusou", "Pediu para não receber mais mensagens — bloqueado"),
        ("interessado", "Respondeu positivamente — ⚠ NECESSITA ATENDIMENTO HUMANO"),
    ]
    for i, (status, desc) in enumerate(legendas, 18):
        cor_f, cor_b = COR_STATUS.get(status, ("000000", "FFFFFF"))
        ws.row_dimensions[i].height = 20

        a = ws.cell(i, 1, status)
        a.fill = PatternFill("solid", fgColor=cor_b)
        a.font = Font(bold=True, size=9, color=cor_f)
        a.alignment = Alignment(vertical="center", indent=1)
        a.border = borda_fina()

        ws.merge_cells(f"B{i}:D{i}")
        b = ws.cell(i, 2, desc)
        b.fill = PatternFill("solid", fgColor=cor_b)
        b.font = Font(size=9, color=cor_f)
        b.alignment = Alignment(vertical="center", indent=1)
        b.border = borda_fina()


# ── Main ───────────────────────────────────────────────────────────────────────
if __name__ == "__main__":
    wb = openpyxl.Workbook()
    criar_aba_leads(wb)
    criar_aba_dashboard(wb)
    wb.save(OUTPUT)
    print(f"✓ Planilha criada: {OUTPUT}")
    print()
    print("Próximos passos:")
    print("  1. Abra leads.xlsx e preencha a aba 'Leads' com seus contatos")
    print("  2. Coluna Telefone: formato 5521999990001 (55 + DDD + número)")
    print("  3. Status inicial de todos os leads: pendente")
    print("  4. Salve e feche o arquivo antes de rodar campanha_whatsapp.py")
