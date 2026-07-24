#!/usr/bin/env python3
"""
CampanhaWhatsApp — Automação de prospecção via Evolution API (WhatsApp gratuito)
Regras: 11-17 contatos/dia | envios entre 10h e 15h30 | Seg-Qui | sem feriados | recontato em 12 dias
"""

import json
import time
import random
import logging
import requests
from datetime import datetime, date, timedelta
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── Configuração de log ────────────────────────────────────────────────────────
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s  %(levelname)-8s  %(message)s",
    handlers=[
        logging.StreamHandler(),
        logging.FileHandler("campanha.log", encoding="utf-8"),
    ],
)
log = logging.getLogger(__name__)

# ── Constantes de status ───────────────────────────────────────────────────────
STATUS_PENDENTE     = "pendente"
STATUS_ABORDADO     = "abordado"
STATUS_IGNORADO     = "ignorado"
STATUS_RECUSOU      = "recusou"
STATUS_INTERESSADO  = "interessado"

# Colunas da aba Leads (índice 1 = coluna A)
COL_NOME            = 1
COL_TELEFONE        = 2
COL_CIDADE          = 3
COL_STATUS          = 4
COL_DATA_CONTATO    = 5
COL_DATA_ULTIMO     = 6
COL_PROXIMA_TENT    = 7
COL_TENTATIVAS      = 8
COL_ULTIMA_RESP     = 9
COL_HORA_RESP       = 10
COL_OBSERVACOES     = 11


# ── Carregamento de configuração ───────────────────────────────────────────────
def carregar_config(caminho="config.json") -> dict:
    with open(caminho, encoding="utf-8") as f:
        return json.load(f)


# ── Verificação de dia útil ────────────────────────────────────────────────
NOMES_DIA = ["Segunda", "Terça", "Quarta", "Quinta", "Sexta", "Sábado", "Domingo"]

# Feriados RJ estaduais e Teresópolis municipais (fixos, complementam a BrasilAPI)
FERIADOS_EXTRAS = {
    "04-23",  # São Jorge — feriado estadual RJ
    "07-06",  # Aniversário de Teresópolis — feriado municipal
}

# Feriados nacionais de fallback (caso BrasilAPI esteja offline)
FERIADOS_NACIONAIS_FIXOS = [
    "01-01", "04-21", "05-01", "09-07",
    "10-12", "11-02", "11-15", "11-20", "12-25",
]


def buscar_feriados_nacionais(ano: int) -> set[str]:
    """Consulta BrasilAPI. Retorna set de datas 'MM-DD'. Fallback em caso de erro."""
    try:
        resp = requests.get(
            f"https://brasilapi.com.br/api/feriados/v1/{ano}",
            timeout=8,
        )
        resp.raise_for_status()
        return {f["date"][5:] for f in resp.json()}  # "YYYY-MM-DD" → "MM-DD"
    except Exception as e:
        log.warning(f"BrasilAPI indisponível ({e}). Usando lista de feriados fixos.")
        return set(FERIADOS_NACIONAIS_FIXOS)


def dia_liberado_para_envio() -> bool:
    """
    Retorna True apenas se hoje for Seg–Qui E não for feriado.
    Feriados verificados: nacionais (BrasilAPI) + estaduais RJ + municipais Teresópolis.
    """
    hoje = date.today()

    # Seg=0 Ter=1 Qua=2 Qui=3 Sex=4 Sáb=5 Dom=6
    if hoje.weekday() > 3:
        log.info(f"Hoje é {NOMES_DIA[hoje.weekday()]} — envios apenas de Segunda a Quinta.")
        return False

    todos_feriados = buscar_feriados_nacionais(hoje.year) | FERIADOS_EXTRAS
    chave = hoje.strftime("%m-%d")  # "MM-DD"

    if chave in todos_feriados:
        log.info(f"Hoje ({hoje.strftime('%d/%m/%Y')}) é feriado — nenhum envio será realizado.")
        return False

    return True


# ── Mapa de id de privacidade (@lid) ──────────────────────────────────────────
# O WhatsApp novo identifica alguns contatos por um "@lid" em vez do numero.
# Ao enviar, guardamos <lid> -> <telefone> para o webhook reconhecer as respostas.
ARQUIVO_LID_MAP = "lid_map.json"


def registrar_lid(telefone: str, data: dict):
    """Extrai o id de privacidade (@lid) da resposta do envio e salva no mapa."""
    try:
        msg_id = data.get("id")
        remote = msg_id.get("remote", "") if isinstance(msg_id, dict) else ""
        if not remote:
            return
        lid = remote.split("@")[0]
        if not lid or lid == telefone:
            return
        mapa = {}
        if Path(ARQUIVO_LID_MAP).exists():
            try:
                mapa = json.loads(Path(ARQUIVO_LID_MAP).read_text(encoding="utf-8"))
            except Exception:
                mapa = {}
        if mapa.get(lid) != telefone:
            mapa[lid] = telefone
            Path(ARQUIVO_LID_MAP).write_text(
                json.dumps(mapa, ensure_ascii=False, indent=2), encoding="utf-8"
            )
            log.info(f"  (id de privacidade {lid} ligado a {telefone})")
    except Exception as e:
        log.warning(f"  Nao consegui registrar o id de privacidade: {e}")


# ── Envio via WAHA (WhatsApp HTTP API) ────────────────────────────────────────
def enviar_mensagem(telefone: str, texto: str, cfg: dict) -> bool:
    """
    Envia mensagem de texto via WAHA.
    O número deve estar no formato: 5521999990001 (55 + DDD + número).
    """
    base_url = cfg["waha"]["url"].rstrip("/")
    sessao   = cfg["waha"]["sessao"]
    api_key  = cfg["waha"].get("api_key", "")

    url = f"{base_url}/api/sendText"
    headers = {"Content-Type": "application/json"}
    if api_key:
        headers["X-Api-Key"] = api_key

    payload = {
        "session": sessao,
        "chatId":  f"{telefone}@c.us",
        "text":    texto,
    }
    try:
        resp = requests.post(url, headers=headers, json=payload, timeout=15)
        resp.raise_for_status()
        # Envio real bem-sucedido retorna um JSON com o id da mensagem.
        # Corpo vazio = o motor do WAHA falhou (ex: getChatById quebrado) e a
        # mensagem NAO foi entregue — tratamos como falha (loud) para nao marcar
        # o lead como abordado sem ter enviado nada.
        try:
            data = resp.json()
        except ValueError:
            log.error(f"  ✗ WAHA respondeu vazio para {telefone} — mensagem NAO enviada "
                      f"(motor do WAHA com problema). Lead sera tentado de novo.")
            return False
        registrar_lid(telefone, data)
        msg_id = data.get("id", "?")
        log.info(f"  ✓ Enviado para {telefone} | msg_id={msg_id}")
        return True
    except requests.HTTPError as e:
        log.error(f"  ✗ Erro HTTP ao enviar para {telefone}: {e.response.text}")
        return False
    except Exception as e:
        log.error(f"  ✗ Erro ao enviar para {telefone}: {e}")
        return False


# ── Planilha Excel ─────────────────────────────────────────────────────────────
def abrir_planilha(caminho: str) -> openpyxl.Workbook:
    if not Path(caminho).exists():
        log.error(f"Arquivo '{caminho}' não encontrado. Execute 'criar_planilha.py' primeiro.")
        raise FileNotFoundError(caminho)
    return openpyxl.load_workbook(caminho)


def salvar_planilha(wb: openpyxl.Workbook, caminho: str):
    wb.save(caminho)
    log.info(f"Planilha salva → {caminho}")


def ler_leads(ws) -> list[dict]:
    """Lê todas as linhas de leads a partir da linha 2."""
    leads = []
    for row in ws.iter_rows(min_row=2, values_only=False):
        nome = row[COL_NOME - 1].value
        if not nome:
            continue
        leads.append({
            "row":        row[0].row,
            "nome":       str(nome).strip(),
            "telefone":   str(int(float(row[COL_TELEFONE - 1].value))) if row[COL_TELEFONE - 1].value else "",
            "cidade":     str(row[COL_CIDADE - 1].value or "").strip(),
            "status":     str(row[COL_STATUS - 1].value or STATUS_PENDENTE).strip(),
            "tentativas": int(row[COL_TENTATIVAS - 1].value or 0),
            "proxima":    row[COL_PROXIMA_TENT - 1].value,
        })
    return leads


def atualizar_linha(ws, row_num: int, status: str, tentativas: int,
                    data_contato_inicial=None, mensagem_tipo="inicial"):
    hoje = date.today()
    ws.cell(row_num, COL_STATUS).value        = status
    ws.cell(row_num, COL_DATA_ULTIMO).value   = hoje.isoformat()
    ws.cell(row_num, COL_TENTATIVAS).value    = tentativas
    if data_contato_inicial:
        ws.cell(row_num, COL_DATA_CONTATO).value = data_contato_inicial.isoformat()
    # Próxima tentativa: vazio se recusou/interessado, +12 dias se abordado
    if status == STATUS_ABORDADO:
        proxima = (hoje + timedelta(days=12)).isoformat()
        ws.cell(row_num, COL_PROXIMA_TENT).value = proxima


def atualizar_dashboard(wb: openpyxl.Workbook):
    """Recalcula os contadores do Dashboard baseado nos dados atuais."""
    ws_leads = wb["Leads"]
    ws_dash  = wb["Dashboard"]

    contadores = {
        STATUS_PENDENTE:    0,
        STATUS_ABORDADO:    0,
        STATUS_IGNORADO:    0,
        STATUS_RECUSOU:     0,
        STATUS_INTERESSADO: 0,
    }
    total_abordados = 0
    abordados_hoje  = 0
    hoje_str = date.today().isoformat()

    for row in ws_leads.iter_rows(min_row=2, values_only=True):
        nome   = row[COL_NOME - 1]
        status = str(row[COL_STATUS - 1] or STATUS_PENDENTE)
        ult    = str(row[COL_DATA_ULTIMO - 1] or "")
        if not nome:
            continue
        if status in contadores:
            contadores[status] += 1
        if status != STATUS_PENDENTE:
            total_abordados += 1
        if ult == hoje_str:
            abordados_hoje += 1

    total = sum(contadores.values())

    # Atualiza células de valor no dashboard (mapeadas por rótulo na coluna A)
    for row in ws_dash.iter_rows(min_row=1, values_only=False):
        label = str(row[0].value or "").strip().lower()
        if not label or len(row) < 2:
            continue
        cel = row[1]
        if "total leads"       in label: cel.value = total
        elif "abordados hoje"  in label: cel.value = abordados_hoje
        elif "total abordados" in label: cel.value = total_abordados
        elif "pendentes"       in label: cel.value = contadores[STATUS_PENDENTE]
        elif "ignorados"       in label: cel.value = contadores[STATUS_IGNORADO]
        elif "recusaram"       in label: cel.value = contadores[STATUS_RECUSOU]
        elif "interessados"    in label: cel.value = contadores[STATUS_INTERESSADO]
        elif "última atualiz"  in label: cel.value = datetime.now().strftime("%d/%m/%Y %H:%M")
        elif "taxa resposta"   in label:
            if total_abordados > 0:
                taxa = (contadores[STATUS_INTERESSADO] / total_abordados) * 100
                cel.value = round(taxa, 1)


# ── Seleção de leads elegíveis ─────────────────────────────────────────────────
def selecionar_leads_do_dia(leads: list[dict], cfg: dict) -> list[dict]:
    hoje = date.today()
    lim_min = cfg["campanha"]["limite_min_por_dia"]
    lim_max = cfg["campanha"]["limite_max_por_dia"]
    max_tent = cfg["campanha"]["max_tentativas"]
    elegíveis = []

    for lead in leads:
        status = lead["status"]
        tent   = lead["tentativas"]

        # Bloqueados permanentemente
        if status in (STATUS_RECUSOU, STATUS_INTERESSADO):
            continue

        # Excede tentativas máximas
        if tent >= max_tent:
            # Marca como ignorado se ainda estiver como "abordado"
            continue

        # Pendente (nunca contactado)
        if status == STATUS_PENDENTE:
            elegíveis.append(lead)
            continue

        # Abordado mas ignorado — verifica se já passaram os dias de recontato
        if status in (STATUS_ABORDADO, STATUS_IGNORADO):
            proxima_raw = lead["proxima"]
            if proxima_raw:
                try:
                    if isinstance(proxima_raw, str):
                        proxima = date.fromisoformat(proxima_raw)
                    elif isinstance(proxima_raw, (date, datetime)):
                        proxima = proxima_raw if isinstance(proxima_raw, date) else proxima_raw.date()
                    else:
                        proxima = hoje  # força recontato se formato inválido
                    if hoje >= proxima:
                        elegíveis.append(lead)
                except ValueError:
                    elegíveis.append(lead)
            else:
                elegíveis.append(lead)

    if not elegíveis:
        return []

    # Modo teste: limita a quantidade, ignorando o mínimo/máximo diário
    mt = cfg.get("modo_teste", {})
    if mt.get("ativo", False):
        max_teste = mt.get("max_envios", 2)
        quantidade = min(max_teste, len(elegíveis))
        selecionados = random.sample(elegíveis, quantidade)
        log.info(f"[MODO TESTE] Elegíveis: {len(elegíveis)} | "
                 f"Selecionados (limite teste {max_teste}): {len(selecionados)}")
        return selecionados

    quantidade = random.randint(min(lim_min, len(elegíveis)), min(lim_max, len(elegíveis)))
    selecionados = random.sample(elegíveis, quantidade)
    log.info(f"Elegíveis: {len(elegíveis)} | Selecionados hoje: {len(selecionados)}")
    return selecionados


# ── Agendamento de horários de envio ──────────────────────────────────────────
def calcular_horarios(quantidade: int, cfg: dict) -> list[datetime]:
    """Distribui os horários de envio aleatoriamente dentro da janela permitida."""
    h_ini_str = cfg["campanha"]["hora_inicio_envio"]
    h_fim_str = cfg["campanha"]["hora_fim_envio"]

    hoje = date.today()
    h_ini = datetime.combine(hoje, datetime.strptime(h_ini_str, "%H:%M").time())
    h_fim = datetime.combine(hoje, datetime.strptime(h_fim_str, "%H:%M").time())

    janela_seg = int((h_fim - h_ini).total_seconds())
    if quantidade <= 1:
        return [h_ini + timedelta(seconds=janela_seg // 2)]

    # Gera timestamps aleatórios, ordenados
    timestamps = sorted(
        [h_ini + timedelta(seconds=random.randint(0, janela_seg)) for _ in range(quantidade)]
    )
    return timestamps


# ── Montagem do texto da mensagem ──────────────────────────────────────────────
def montar_mensagem(lead: dict, cfg: dict) -> str:
    if lead["tentativas"] == 0:
        template = cfg["mensagem"]["inicial"]
    else:
        template = cfg["mensagem"]["recontato"]
    return template.format(nome=lead["nome"].split()[0], cidade=lead["cidade"])


# ── Loop principal ────────────────────────────────────────────────────────────
def executar_campanha():
    cfg = carregar_config()
    arquivo = cfg["campanha"]["arquivo_leads"]

    log.info("=" * 60)
    log.info(f"  Campanha WhatsApp — {date.today().strftime('%d/%m/%Y')} ({NOMES_DIA[date.today().weekday()]})")
    log.info("=" * 60)

    # Modo teste: ignora travas de dia/feriado e limita envios (config.json -> "modo_teste")
    mt = cfg.get("modo_teste", {})
    modo_teste = mt.get("ativo", False)

    if modo_teste and mt.get("ignorar_dia_horario", True):
        log.warning("=" * 60)
        log.warning(" >>> MODO TESTE ATIVO — trava de dia/feriado IGNORADA <<<")
        log.warning(f" Envios limitados a {mt.get('max_envios', 2)} | sem espera de horario")
        log.warning("=" * 60)
    else:
        if not dia_liberado_para_envio():
            log.info("Encerrando sem envios.")
            return
        # Trava de horario: nunca enviar depois do fim da janela de envio.
        # Protege contra execucoes tardias (ex: a tarefa "executar se perdido"
        # disparando a noite) — que senao mandariam todos os horarios ja vencidos
        # de uma vez (rajada). Fora da janela, encerra sem enviar; roda no proximo dia.
        h_fim = datetime.strptime(cfg["campanha"]["hora_fim_envio"], "%H:%M").time()
        if datetime.now().time() > h_fim:
            log.info(f"Agora e depois de {cfg['campanha']['hora_fim_envio']} — fora da "
                     f"janela de envio. Encerrando sem envios (roda amanha no horario).")
            return

    wb = abrir_planilha(arquivo)
    ws = wb["Leads"]
    leads = ler_leads(ws)

    # Marca como "ignorado" leads abordados há mais de 12 dias sem resposta
    for lead in leads:
        if lead["status"] == STATUS_ABORDADO:
            proxima_raw = lead["proxima"]
            if proxima_raw:
                try:
                    if isinstance(proxima_raw, str):
                        proxima = date.fromisoformat(proxima_raw)
                    elif isinstance(proxima_raw, datetime):
                        proxima = proxima_raw.date()
                    else:
                        proxima = proxima_raw
                    if date.today() >= proxima and lead["tentativas"] < cfg["campanha"]["max_tentativas"]:
                        ws.cell(lead["row"], COL_STATUS).value = STATUS_IGNORADO
                        lead["status"] = STATUS_IGNORADO
                except Exception:
                    pass

    selecionados = selecionar_leads_do_dia(leads, cfg)
    if not selecionados:
        log.info("Nenhum lead elegivel para hoje. Encerrando.")
        salvar_planilha(wb, arquivo)
        return

    horarios = calcular_horarios(len(selecionados), cfg)

    log.info(f"Envios programados para hoje: {len(selecionados)}")
    for lead, horario in zip(selecionados, horarios):
        agora = datetime.now()
        if not modo_teste and horario > agora:   # em modo teste, envia na hora
            espera = (horario - agora).total_seconds()
            log.info(f"  Aguardando {int(espera)}s ate {horario.strftime('%H:%M:%S')} para enviar a {lead['nome']}...")
            time.sleep(espera)

        mensagem = montar_mensagem(lead, cfg)
        log.info(f"Enviando para {lead['nome']} ({lead['telefone']})...")
        sucesso = enviar_mensagem(lead["telefone"], mensagem, cfg)

        if sucesso:
            tentativas = lead["tentativas"] + 1
            data_inicial = date.fromisoformat(str(ws.cell(lead["row"], COL_DATA_CONTATO).value)) \
                if ws.cell(lead["row"], COL_DATA_CONTATO).value else date.today()
            atualizar_linha(ws, lead["row"], STATUS_ABORDADO, tentativas,
                            data_contato_inicial=data_inicial if tentativas == 1 else None)
            atualizar_dashboard(wb)
            salvar_planilha(wb, arquivo)
        else:
            log.warning(f"  Falha ao enviar para {lead['nome']} — sera tentado novamente na proxima execucao.")

    log.info("=" * 60)
    log.info("  Campanha finalizada.")
    log.info("=" * 60)


if __name__ == "__main__":
    executar_campanha()
