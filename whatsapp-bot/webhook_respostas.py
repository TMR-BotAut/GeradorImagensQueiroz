#!/usr/bin/env python3
"""
WebhookRespostas -- Servidor WhatsApp via WAHA
- Leads conhecidos (na planilha): FAQ + classificacao de resposta
- Contatos desconhecidos: menu interativo com triagem por produto
"""

import json
import logging
import re
import time
from datetime import datetime, date
from pathlib import Path

import requests as http_requests
from flask import Flask, request, jsonify
import openpyxl

# -- Log -----------------------------------------------------------------------
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s %(levelname)-8s %(message)s",
    handlers=[
        logging.StreamHandler(),
        logging.FileHandler("webhook.log", encoding="utf-8"),
    ],
)
log = logging.getLogger(__name__)

# -- Config --------------------------------------------------------------------
with open("config.json", encoding="utf-8") as f:
    CFG = json.load(f)

ARQUIVO_LEADS = CFG["campanha"]["arquivo_leads"]
WAHA_URL = CFG["waha"]["url"].rstrip("/")
WAHA_SESSAO = CFG["waha"]["sessao"]
WAHA_API_KEY = CFG["waha"].get("api_key", "")

# Mapa <id de privacidade @lid> -> <telefone real>, alimentado pela campanha.
# O WhatsApp novo identifica alguns contatos por um "@lid" em vez do numero;
# este mapa permite reconhecer o lead quando a resposta chega.
ARQUIVO_LID_MAP = "lid_map.json"

# Rodrigo -- parceiro para produtos fora de saude
RODRIGO_NOME = "Rodrigo"
RODRIGO_TELEFONE = "5521988541324"

# Numero que recebe alertas em tempo real (config.json -> "alertas": {"numero": "55..."})
ALERTA_TELEFONE = str(CFG.get("alertas", {}).get("numero", "") or "")

# Status
STATUS_ABORDADO = "abordado"
STATUS_RECUSOU = "recusou"
STATUS_INTERESSADO = "interessado"

# Colunas da aba Leads
COL_NOME = 1
COL_TELEFONE = 2
COL_CIDADE = 3
COL_STATUS = 4
COL_TENTATIVAS = 8
COL_ULTIMA_R = 9
COL_HORA_R = 10
COL_OBS = 11

# Menus
MENU_CATEGORIA = (
    "O que você está buscando, {apelido}? 😊\n\n"
    "1️⃣ *Plano de Saúde e Odontológico*\n"
    "2️⃣ *Seguros* (Veículo, Vida, Residência, Equipamentos, Viagem)\n"
    "3️⃣ *Soluções Financeiras* (Consórcio, Financiamento, Previdência, Seguros)\n\n"
    "Responda com o número da opção."
)

MENU_SEGUROS = (
    "Entendi! Vamos lá com os Seguros. Qual te interessa? (responde o número)\n\n"
    "1️⃣ Seguro de Veículo\n"
    "2️⃣ Seguro de Vida\n"
    "3️⃣ Seguro Residencial\n"
    "4️⃣ Seguro de Equipamentos Portáteis\n"
    "5️⃣ Seguro de Viagem\n"
    "6️⃣ Seguro de Responsabilidade Civil\n"
    "7️⃣ Seguro Transporte / Carga\n"
    "8️⃣ Seguro Pet"
)

MENU_FINANCEIRO = (
    "Entendi! Vamos lá com Soluções Financeiras. Qual te interessa? (responde o número)\n\n"
    "1️⃣ Consórcio\n"
    "2️⃣ Financiamento\n"
    "3️⃣ Previdência\n"
    "4️⃣ Responsabilidade Civil\n"
    "5️⃣ Seguro de Vida"
)

PRODUTOS_SEGUROS = {
    "1": "Seguro de Veículo",
    "2": "Seguro de Vida",
    "3": "Seguro Residencial",
    "4": "Seguro de Equipamentos Portáteis",
    "5": "Seguro de Viagem",
    "6": "Seguro de Responsabilidade Civil",
    "7": "Seguro Transporte / Carga",
    "8": "Seguro Pet",
}

PRODUTOS_FINANCEIRO = {
    "1": "Consórcio",
    "2": "Financiamento",
    "3": "Previdência",
    "4": "Responsabilidade Civil",
    "5": "Seguro de Vida",
}

# Controle de mensagens ja processadas (evita duplicatas)
PROCESSADOS: set = set()

# Estado das conversas de contatos desconhecidos
CONVERSAS: dict = {}

app = Flask(__name__)


# -- Resolucao do id de privacidade (@lid) -------------------------------------
def resolver_telefone(from_jid: str) -> str:
    """Converte o id do remetente no telefone real.

    Para @c.us o proprio id ja e o telefone. Para @lid (id de privacidade),
    tenta o lid_map.json e, se nao achar, consulta o WAHA (/api/contacts);
    o resultado e salvo no lid_map.json para as proximas mensagens.
    """
    raw_id = from_jid.split("@")[0]
    if not from_jid.endswith("@lid"):
        return raw_id

    mapa = {}
    try:
        if Path(ARQUIVO_LID_MAP).exists():
            mapa = json.loads(Path(ARQUIVO_LID_MAP).read_text(encoding="utf-8"))
            if raw_id in mapa:
                return str(mapa[raw_id])
    except Exception as e:
        log.warning(f"Nao consegui ler {ARQUIVO_LID_MAP}: {e}")

    try:
        headers = {"X-Api-Key": WAHA_API_KEY} if WAHA_API_KEY else {}
        resp = http_requests.get(
            f"{WAHA_URL}/api/contacts",
            params={"contactId": f"{raw_id}@lid", "session": WAHA_SESSAO},
            headers=headers,
            timeout=10,
        )
        if resp.ok:
            numero = (resp.json() or {}).get("number")
            if numero:
                mapa[raw_id] = str(numero)
                try:
                    Path(ARQUIVO_LID_MAP).write_text(
                        json.dumps(mapa, indent=2, ensure_ascii=False),
                        encoding="utf-8",
                    )
                except Exception as e:
                    log.warning(f"Nao consegui salvar {ARQUIVO_LID_MAP}: {e}")
                log.info(f" @lid {raw_id} resolvido via WAHA -> {numero}")
                return str(numero)
    except Exception as e:
        log.warning(f" Falha ao resolver @lid {raw_id} via WAHA: {e}")

    return raw_id


# -- Envio de mensagens --------------------------------------------------------
def enviar_resposta(telefone: str, texto: str) -> bool:
    url = f"{WAHA_URL}/api/sendText"
    headers = {"Content-Type": "application/json"}
    if WAHA_API_KEY:
        headers["X-Api-Key"] = WAHA_API_KEY
    # Usa o chatId original se ja tiver @, senao adiciona @c.us
    chat_id = telefone if "@" in telefone else f"{telefone}@c.us"
    payload = {"session": WAHA_SESSAO, "chatId": chat_id, "text": texto}
    try:
        time.sleep(2)
        resp = http_requests.post(url, headers=headers, json=payload, timeout=15)
        resp.raise_for_status()
        log.info(f" -> Mensagem enviada para {telefone}")
        return True
    except Exception as e:
        log.error(f" ERRO ao enviar para {telefone}: {e}")
        return False


def notificar_rodrigo(apelido: str, telefone_cliente: str, produto: str, telefone_informado: str = None):
    msg = (
        f"🔔 Contato novo - {apelido}\n"
        f"📱 Telefone (WhatsApp): {telefone_cliente}\n"
    )
    if telefone_informado and telefone_informado != telefone_cliente:
        msg += f"📱 Telefone informado: {telefone_informado}\n"
    msg += f"🛡️ Interesse: {produto}"

    enviar_resposta(RODRIGO_TELEFONE, msg)
    log.info(f" Rodrigo notificado: {apelido} / {produto}")


def alertar_dono(texto: str):
    """Alerta em tempo real para o numero em config alertas.numero (silencioso se vazio)."""
    if ALERTA_TELEFONE:
        enviar_resposta(ALERTA_TELEFONE, texto)


# -- Comandos administrativos (dono manda mensagem para o bot) ------------------
# Numeros autorizados a enviar comandos administrativos (ex: "bloquear 5521...").
# Inclui o numero de alertas (dono) e o Rodrigo (parceiro).
NUMEROS_ADMIN = [n for n in (ALERTA_TELEFONE, RODRIGO_TELEFONE) if n]


def eh_admin(telefone: str) -> bool:
    """True se o remetente e um dos numeros autorizados (dono ou Rodrigo)."""
    tel = "".join(filter(str.isdigit, telefone))
    if not tel:
        return False
    for adm in NUMEROS_ADMIN:
        admd = "".join(filter(str.isdigit, adm))
        if admd and (tel == admd or tel[-11:] == admd[-11:]):
            return True
    return False


def bloquear_lead_por_comando(numero: str):
    """Marca o lead como recusou (nunca mais recebe campanha).

    Retorna ("ok", nome), ("nao_encontrado", None) ou ("erro", detalhe).
    """
    if not Path(ARQUIVO_LEADS).exists():
        return ("erro", f"arquivo {ARQUIVO_LEADS} nao encontrado")
    alvo = "".join(filter(str.isdigit, numero))
    try:
        wb = openpyxl.load_workbook(ARQUIVO_LEADS)
        ws = wb["Leads"]
        for row in ws.iter_rows(min_row=2):
            tel = "".join(filter(str.isdigit, str(row[COL_TELEFONE - 1].value or "")))
            if not tel:
                continue
            if tel == alvo or tel[-11:] == alvo[-11:]:
                row_num = row[0].row
                nome = str(row[COL_NOME - 1].value or "?")
                ws.cell(row_num, COL_STATUS).value = STATUS_RECUSOU
                ws.cell(row_num, COL_TENTATIVAS).value = 99
                obs = str(ws.cell(row_num, COL_OBS).value or "")
                if "NAO PERTURBE" not in obs:
                    marca = f"[NAO PERTURBE {date.today().strftime('%d/%m/%Y')}]"
                    ws.cell(row_num, COL_OBS).value = (obs + " " + marca).strip()
                atualizar_dashboard(wb)
                wb.save(ARQUIVO_LEADS)
                return ("ok", nome)
        return ("nao_encontrado", None)
    except Exception as e:
        log.error(f" Erro ao bloquear por comando: {e}")
        return ("erro", str(e))


def processar_comando_admin(from_jid: str, texto: str) -> bool:
    """Processa comandos do dono. Retorna True se o texto era um comando.

    Comandos:
        bloquear 5521966628839          -> marca recusou na planilha
        bloquear 5521966628839 avisar   -> idem + avisa a pessoa do descadastro
    """
    partes = texto.strip().lower().split()
    if not partes or partes[0] != "bloquear":
        return False

    digitos = "".join(ch for ch in texto if ch.isdigit())
    if len(digitos) < 10:
        enviar_resposta(
            from_jid,
            "Uso: *bloquear 5521966628839*\n"
            "Para também avisar a pessoa: *bloquear 5521966628839 avisar*",
        )
        return True

    resultado, info = bloquear_lead_por_comando(digitos)

    if resultado == "nao_encontrado":
        enviar_resposta(from_jid, f"Número {digitos} não encontrado na planilha — nada foi alterado.")
        return True

    if resultado == "erro":
        enviar_resposta(
            from_jid,
            f"⚠️ Não consegui salvar a planilha (está aberta no Excel?).\nDetalhe: {info[:120]}",
        )
        return True

    log.info(f" COMANDO ADMIN: bloquear {digitos} -> {info}")
    if "avisar" in partes:
        msg_opt_out = CFG["mensagem"].get(
            "opt_out_confirmacao",
            "Entendido! Você não receberá mais nossas mensagens por aqui. Obrigado! 😊",
        )
        enviar_resposta(digitos, msg_opt_out)
        enviar_resposta(from_jid, f"✓ Bloqueado: {info} ({digitos}) — pessoa avisada do descadastro.")
    else:
        enviar_resposta(from_jid, f"✓ Bloqueado: {info} ({digitos})")
    return True


# -- Fluxo de contatos desconhecidos -------------------------------------------
def processar_desconhecido(from_jid: str, telefone: str, texto: str):
    """
    from_jid: JID completo para envio (ex: 28527491567699@lid ou 5521999@c.us)
    telefone: numero puro para chave do dicionario e planilha
    """
    estado = CONVERSAS.get(telefone, {"etapa": "inicio"})
    etapa = estado.get("etapa", "inicio")
    texto = texto.strip()

    if etapa == "inicio":
        CONVERSAS[telefone] = {"etapa": "aguarda_nome", "from_jid": from_jid, "numero_original": telefone}
        # Se o remetente veio como @lid e nao conseguimos resolver o numero real,
        # nao faz sentido pedir confirmacao dele — pedimos o numero diretamente.
        lid_sem_numero = from_jid.endswith("@lid") and telefone == from_jid.split("@")[0]
        if lid_sem_numero:
            saudacao = CFG["mensagem"].get(
                "triagem_saudacao_sem_numero",
                "Olá! Como vai?! 😊\n\nMeu nome é [SEU NOME], trabalho com seguros e planos de saúde.\n\nMe conta, qual é o seu nome? E qual o melhor número para falar com você?"
            )
        else:
            saudacao = CFG["mensagem"].get("triagem_saudacao", "Olá! Como vai?! 😊\n\nMeu nome é [SEU NOME], trabalho com seguros e planos de saúde.\n\nMe conta, qual é o seu nome? Pode confirmar se {number} é seu contato preferencial, ou se prefere informar outro?")
            saudacao = saudacao.replace("{number}", telefone)
        enviar_resposta(from_jid, saudacao)
        log.info(f" Novo contato {from_jid} - iniciando fluxo de triagem")
        return

    from_jid = estado.get("from_jid", from_jid)

    if etapa == "aguarda_nome":
        # Tenta extrair nome e possível número da resposta
        partes = texto.strip().split()
        if not partes:
            return
        apelido = partes[0].capitalize()

        # Junta os digitos do resto da mensagem (aceita "21 97293-2120", "(21) 97293 2120" etc.)
        digitos_resto = "".join(ch for ch in " ".join(partes[1:]) if ch.isdigit())
        numero_informado = digitos_resto if len(digitos_resto) >= 10 else None

        CONVERSAS[telefone]["apelido"] = apelido

        # Se informou número diferente do original
        if numero_informado and numero_informado[-11:] != str(estado.get("numero_original", ""))[-11:]:
            CONVERSAS[telefone]["numero_informado"] = numero_informado
            CONVERSAS[telefone]["etapa"] = "confirma_numero"
            confirma_msg = CFG["mensagem"].get("triagem_confirma_numero", "Seu melhor número para contato é o *{numero_informado}*?")
            confirma_msg = confirma_msg.replace("{numero_informado}", numero_informado)
            enviar_resposta(from_jid, confirma_msg)
        else:
            # Sem número diferente, segue pra categoria
            CONVERSAS[telefone]["etapa"] = "aguarda_categoria"
            menu = CFG["mensagem"].get("triagem_menu_categoria", MENU_CATEGORIA)
            enviar_resposta(from_jid, menu.format(apelido=apelido))
        return

    if etapa == "confirma_numero":
        apelido = estado.get("apelido", "")
        digitos = "".join(filter(str.isdigit, texto))
        low = texto.lower()

        if len(digitos) >= 10:
            # Informou (ou corrigiu) um número diretamente
            CONVERSAS[telefone]["numero_informado"] = digitos
            CONVERSAS[telefone]["etapa"] = "aguarda_categoria"
            menu = CFG["mensagem"].get("triagem_menu_categoria", MENU_CATEGORIA)
            enviar_resposta(from_jid, menu.format(apelido=apelido))
        elif any(p in low for p in ["não", "nao", "outro", "errado", "corrig"]):
            enviar_resposta(from_jid, f"Sem problema, {apelido}! Qual é o número correto?")
            CONVERSAS[telefone]["etapa"] = "aguarda_numero_corrigido"
        elif any(p in low for p in ["sim", "isso", "ok", "confirmo", "correto", "certo", "pode ser", "esse mesmo", "é esse", "e esse"]):
            CONVERSAS[telefone]["etapa"] = "aguarda_categoria"
            menu = CFG["mensagem"].get("triagem_menu_categoria", MENU_CATEGORIA)
            enviar_resposta(from_jid, menu.format(apelido=apelido))
        else:
            enviar_resposta(from_jid, f"Por favor, {apelido}, me confirma se o número está certo ou me passa o número correto. 😊")
        return

    if etapa == "aguarda_numero_corrigido":
        apelido = estado.get("apelido", "")
        digitos = "".join(filter(str.isdigit, texto))
        if len(digitos) >= 10:
            CONVERSAS[telefone]["numero_informado"] = digitos
            CONVERSAS[telefone]["etapa"] = "aguarda_categoria"
            menu = CFG["mensagem"].get("triagem_menu_categoria", MENU_CATEGORIA)
            enviar_resposta(from_jid, menu.format(apelido=apelido))
        else:
            enviar_resposta(from_jid, f"Por favor, {apelido}, informe um número válido (ex: 5521987654321).")
        return

    if etapa == "aguarda_categoria":
        apelido = estado.get("apelido", "")
        if texto == "1":
            # Plano de Saude e Odontologico -> atendimento interno
            msg_saude = CFG["mensagem"].get("triagem_saude_interno", "Ótima escolha, {apelido}! 😊\n\nNossa equipe especializada em *Plano de Saúde e Odontológico* vai entrar em contato com você em breve.\n\nFique à vontade para perguntar qualquer coisa por aqui!")
            enviar_resposta(from_jid, msg_saude.format(apelido=apelido))
            registrar_lead_desconhecido(telefone, estado, "Plano de Saúde e Odontológico", "saude_interno")
            log.info(f" {apelido} ({telefone}) - Saude/Odonto -> ATENDIMENTO INTERNO")
            del CONVERSAS[telefone]
        elif texto == "2":
            CONVERSAS[telefone]["categoria"] = "seguros"
            CONVERSAS[telefone]["etapa"] = "aguarda_produto_seguro"
            menu = CFG["mensagem"].get("triagem_menu_seguros", MENU_SEGUROS)
            enviar_resposta(from_jid, menu)
        elif texto == "3":
            CONVERSAS[telefone]["categoria"] = "financeiro"
            CONVERSAS[telefone]["etapa"] = "aguarda_produto_financeiro"
            menu = CFG["mensagem"].get("triagem_menu_financeiro", MENU_FINANCEIRO)
            enviar_resposta(from_jid, menu)
        else:
            enviar_resposta(
                from_jid,
                f"Por favor, {apelido}, responda apenas com *1*, *2* ou *3*. \U0001f60a\n\n"
                + MENU_CATEGORIA.format(apelido=apelido)
            )
        return

    if etapa == "aguarda_produto_seguro":
        apelido = estado.get("apelido", "")
        tel_cliente = telefone
        tel_informado = estado.get("numero_informado", None)
        produto = PRODUTOS_SEGUROS.get(texto)

        if not produto:
            menu = CFG["mensagem"].get("triagem_menu_seguros", MENU_SEGUROS)
            enviar_resposta(
                from_jid,
                f"Por favor, {apelido}, responda com um número de 1 a 8. 😊\n\n" + menu
            )
            return

        msg_encaminhado = CFG["mensagem"].get("triagem_seguros_encaminhado", "*{nome_parceiro}* é especialista em *{produto}* e vai entrar em contato com você em breve, {apelido}! 😊\n\nCaso prefira, pode chamá-lo diretamente: (21) 98854-1324")
        msg_encaminhado = msg_encaminhado.replace("{nome_parceiro}", RODRIGO_NOME).replace("{produto}", produto).replace("{apelido}", apelido)
        enviar_resposta(from_jid, msg_encaminhado)

        notificar_rodrigo(apelido, tel_cliente, produto, tel_informado)
        registrar_lead_desconhecido(telefone, estado, produto, "rodrigo")
        del CONVERSAS[telefone]
        return

    if etapa == "aguarda_produto_financeiro":
        apelido = estado.get("apelido", "")
        tel_cliente = telefone
        tel_informado = estado.get("numero_informado", None)
        produto = PRODUTOS_FINANCEIRO.get(texto)

        if not produto:
            menu = CFG["mensagem"].get("triagem_menu_financeiro", MENU_FINANCEIRO)
            enviar_resposta(
                from_jid,
                f"Por favor, {apelido}, responda com um número de 1 a 5. 😊\n\n" + menu
            )
            return

        msg_encaminhado = CFG["mensagem"].get("triagem_financeiro_encaminhado", "*{nome_parceiro}* é especialista em *{produto}* e vai entrar em contato com você em breve, {apelido}! 😊\n\nCaso prefira, pode chamá-lo diretamente: (21) 98854-1324")
        msg_encaminhado = msg_encaminhado.replace("{nome_parceiro}", RODRIGO_NOME).replace("{produto}", produto).replace("{apelido}", apelido)
        enviar_resposta(from_jid, msg_encaminhado)

        notificar_rodrigo(apelido, tel_cliente, produto, tel_informado)
        registrar_lead_desconhecido(telefone, estado, produto, "rodrigo")
        del CONVERSAS[telefone]
        return


def registrar_lead_desconhecido(telefone: str, estado: dict, produto: str, destino: str):
    if not Path(ARQUIVO_LEADS).exists():
        return
    try:
        apelido = estado.get("apelido", "")
        wb = openpyxl.load_workbook(ARQUIVO_LEADS)
        ws = wb["Leads"]
        nova_linha = ws.max_row + 1
        ws.cell(nova_linha, COL_NOME).value = apelido
        ws.cell(nova_linha, COL_TELEFONE).value = telefone
        ws.cell(nova_linha, COL_STATUS).value = STATUS_INTERESSADO
        ws.cell(nova_linha, COL_HORA_R).value = datetime.now().strftime("%d/%m/%Y %H:%M")
        obs = f"[INBOUND] Produto: {produto} | Tel: {telefone} | Encaminhado: {destino.upper()}"
        ws.cell(nova_linha, COL_OBS).value = obs
        atualizar_dashboard(wb)
        wb.save(ARQUIVO_LEADS)
        log.info(f" Lead {apelido} ({telefone}) registrado -> {produto}")
    except Exception as e:
        log.error(f" Erro ao registrar lead desconhecido: {e}")


# -- Fluxo de leads conhecidos -------------------------------------------------
def verificar_faq(texto: str):
    texto_lower = texto.lower().strip()
    faq = CFG.get("faq", {})
    for tema, dados in faq.items():
        if tema.startswith("_"):
            continue
        for gatilho in dados.get("gatilhos", []):
            if gatilho in texto_lower:
                return tema, dados["resposta"]
    return None, None


def interpretar_opcao_menu(texto: str):
    """Se a resposta for so um numero de 1 a 5 (o menu da campanha), retorna
    'opcao_1'..'opcao_5'. Senao (ex: 'quero auto'), retorna None e cai na
    classificacao normal. Aceita '2', ' 2 ', '2.', '2)', '2-'."""
    m = re.fullmatch(r"\s*([1-5])\s*[\.\)\-]?\s*", texto or "")
    return f"opcao_{m.group(1)}" if m else None


def classificar_resposta(texto: str) -> str:
    texto_lower = texto.lower().strip()
    for palavra in CFG["palavras_recusa"]:
        if palavra in texto_lower:
            return STATUS_RECUSOU
    tema_faq, _ = verificar_faq(texto)
    if tema_faq:
        return "faq"
    for palavra in CFG["palavras_interesse"]:
        if palavra in texto_lower:
            return STATUS_INTERESSADO
    return "aguarda_humano"


def atualizar_lead(from_jid: str, telefone: str, texto: str, classificacao: str):
    # from_jid = destino para responder (pode ser @lid); telefone = numero real (busca na planilha)
    if not Path(ARQUIVO_LEADS).exists():
        log.warning(f"Arquivo {ARQUIVO_LEADS} nao encontrado.")
        return False

    telefone_limpo = "".join(filter(str.isdigit, telefone))

    try:
        wb = openpyxl.load_workbook(ARQUIVO_LEADS)
        ws = wb["Leads"]
        encontrado = False

        for row in ws.iter_rows(min_row=2, values_only=False):
            tel_planilha = str(row[COL_TELEFONE - 1].value or "")
            tel_limpo = "".join(filter(str.isdigit, tel_planilha))

            if tel_limpo[-11:] == telefone_limpo[-11:] or tel_limpo == telefone_limpo:
                row_num = row[0].row
                nome = row[COL_NOME - 1].value or "?"
                apelido = str(nome).split()[0] if nome else ""

                if classificacao == "opcao_1":
                    # 1 = Plano de saude / odonto -> atendimento interno (dono)
                    novo_status = STATUS_INTERESSADO
                    enviar_resposta(from_jid, CFG["mensagem"].get(
                        "menu_saude",
                        "Perfeito. Sobre plano de saude e odontologico, nossa equipe ja vai te "
                        "atender por aqui. Pode perguntar o que precisar."))
                    log.info(f" MENU 1 (Saude/Odonto) -> {nome} - INTERNO")
                    alertar_dono(
                        f"🔥 *Lead interessado — Saúde/Odonto*\n"
                        f"👤 {nome}\n📱 {tel_planilha}\n\nResponda o quanto antes!")

                elif classificacao in ("opcao_2", "opcao_3", "opcao_4"):
                    # 2/3/4 -> encaminha para o Rodrigo (parceiro)
                    novo_status = STATUS_INTERESSADO
                    produto = {
                        "opcao_2": "Seguro de Auto",
                        "opcao_3": "Residencial / Vida / Equipamentos",
                        "opcao_4": "Ainda indeciso (quer ver opcoes)",
                    }[classificacao]
                    enviar_resposta(from_jid, CFG["mensagem"].get(
                        "menu_rodrigo",
                        "Otimo. Vou te encaminhar para o Rodrigo, nosso especialista, que ja vai te "
                        "atender. Se preferir, fale direto: (21) 98854-1324."))
                    notificar_rodrigo(apelido, tel_planilha, produto)
                    log.info(f" MENU {classificacao[-1]} ({produto}) -> {nome} - RODRIGO")

                elif classificacao == "opcao_5":
                    # 5 = Nao tem interesse -> sai da lista (recusou), de verdade
                    novo_status = STATUS_RECUSOU
                    enviar_resposta(from_jid, CFG["mensagem"].get(
                        "menu_optout",
                        "Tudo bem, sem problema. Ja removi seu contato da nossa lista. Obrigado!"))
                    log.info(f" MENU 5 (Opt-out) -> {nome} - RECUSOU")

                elif classificacao == STATUS_RECUSOU:
                    novo_status = STATUS_RECUSOU
                    log.info(f" RECUSOU -> {nome} ({tel_planilha})")

                elif classificacao == "faq":
                    tema_faq, resposta_faq = verificar_faq(texto)
                    cidade = str(row[COL_CIDADE - 1].value or "sua cidade").strip()
                    resposta_faq = resposta_faq.replace("[CIDADE]", cidade)
                    enviar_resposta(from_jid, resposta_faq)
                    novo_status = STATUS_ABORDADO
                    log.info(f" FAQ '{tema_faq}' respondido para {nome}")

                elif classificacao == STATUS_INTERESSADO:
                    novo_status = STATUS_INTERESSADO
                    log.info(f" INTERESSADO -> {nome} ({tel_planilha}) - ATENCAO HUMANA")
                    alertar_dono(
                        f"🔥 *Lead com INTERESSE!*\n"
                        f"👤 {nome}\n"
                        f"📱 {tel_planilha}\n"
                        f"💬 \"{str(texto)[:150]}\"\n\n"
                        f"Lead quente — responda o quanto antes!"
                    )

                else:
                    novo_status = STATUS_INTERESSADO
                    log.info(f" RESPOSTA -> {nome} ({tel_planilha}) - REVISAR")
                    alertar_dono(
                        f"✋ *Resposta para revisar*\n"
                        f"👤 {nome}\n"
                        f"📱 {tel_planilha}\n"
                        f"💬 \"{str(texto)[:150]}\"\n\n"
                        f"O bot não soube classificar — veja se precisa responder."
                    )

                ws.cell(row_num, COL_STATUS).value = novo_status
                ws.cell(row_num, COL_ULTIMA_R).value = texto[:200]
                ws.cell(row_num, COL_HORA_R).value = datetime.now().strftime("%d/%m/%Y %H:%M")

                if novo_status == STATUS_INTERESSADO:
                    obs_atual = str(ws.cell(row_num, COL_OBS).value or "")
                    flag = f"[AGUARDA ATENDIMENTO {date.today().strftime('%d/%m')}]"
                    if flag not in obs_atual:
                        ws.cell(row_num, COL_OBS).value = f"{flag} {obs_atual}".strip()

                encontrado = True
                break

        if not encontrado:
            return False

        atualizar_dashboard(wb)
        wb.save(ARQUIVO_LEADS)
        log.info(f"Planilha atualizada -> {ARQUIVO_LEADS}")
        return True

    except Exception as e:
        log.error(f"Erro ao atualizar planilha: {e}")
        return False


def lead_existe_na_planilha(telefone: str) -> bool:
    if not Path(ARQUIVO_LEADS).exists():
        return False
    telefone_limpo = "".join(filter(str.isdigit, telefone))
    try:
        wb = openpyxl.load_workbook(ARQUIVO_LEADS, read_only=True)
        ws = wb["Leads"]
        for row in ws.iter_rows(min_row=2, values_only=True):
            tel = "".join(filter(str.isdigit, str(row[COL_TELEFONE - 1] or "")))
            if tel[-11:] == telefone_limpo[-11:] or tel == telefone_limpo:
                return True
    except Exception:
        pass
    return False


def atualizar_dashboard(wb):
    STATUS_PENDENTE = "pendente"
    COL_DATA_ULTIMO = 6
    try:
        ws_leads = wb["Leads"]
        ws_dash = wb["Dashboard"]
    except KeyError:
        return

    contadores = {
        STATUS_PENDENTE: 0, STATUS_ABORDADO: 0,
        "ignorado": 0, STATUS_RECUSOU: 0, STATUS_INTERESSADO: 0,
    }
    total_abordados = 0
    abordados_hoje = 0
    hoje_str = date.today().isoformat()

    for row in ws_leads.iter_rows(min_row=2, values_only=True):
        nome = row[COL_NOME - 1]
        status = str(row[COL_STATUS - 1] or STATUS_PENDENTE)
        ult = str(row[COL_DATA_ULTIMO - 1] or "")
        if not nome:
            continue
        if status in contadores:
            contadores[status] += 1
        if status != STATUS_PENDENTE:
            total_abordados += 1
        if ult == hoje_str:
            abordados_hoje += 1

    total = sum(contadores.values())

    for row in ws_dash.iter_rows(min_row=1, values_only=False):
        label = str(row[0].value or "").strip().lower()
        if not label or len(row) < 2:
            continue
        cel = row[1]
        if "total leads" in label: cel.value = total
        elif "abordados hoje" in label: cel.value = abordados_hoje
        elif "total abordados" in label: cel.value = total_abordados
        elif "pendentes" in label: cel.value = contadores[STATUS_PENDENTE]
        elif "ignorados" in label: cel.value = contadores["ignorado"]
        elif "recusaram" in label: cel.value = contadores[STATUS_RECUSOU]
        elif "interessados" in label: cel.value = contadores[STATUS_INTERESSADO]
        elif "ultima atualiz" in label: cel.value = datetime.now().strftime("%d/%m/%Y %H:%M")
        elif "taxa resposta" in label:
            if total_abordados > 0:
                taxa = (contadores[STATUS_INTERESSADO] / total_abordados) * 100
                cel.value = round(taxa, 1)


# -- Rotas ---------------------------------------------------------------------
@app.route("/webhook", methods=["GET", "POST"])
def receber_mensagem():
    if request.method == "GET":
        return jsonify({"status": "ok"}), 200

    payload = request.get_json(silent=True)
    if not payload:
        return jsonify({"status": "ignored"}), 200

    try:
        evento = payload.get("event", "")
        if evento != "message":
            return jsonify({"status": "ignored"}), 200

        msg = payload.get("payload", {})
        msg_id = msg.get("id", "")
        from_me = msg.get("fromMe", False)
        from_jid = msg.get("from", "") or msg.get("chatId", "")
        body = msg.get("body", "")
        texto = body if isinstance(body, str) else ""

        if from_me:
            # Mensagens da propria conta sao ignoradas, EXCETO comandos administrativos
            # que o dono digita no chat consigo mesmo ("Mensagem para si mesmo").
            if texto.strip().lower().startswith("bloquear"):
                tel_self = resolver_telefone(from_jid)
                if eh_admin(tel_self) and processar_comando_admin(from_jid, texto):
                    return jsonify({"status": "admin_cmd_self"}), 200
            return jsonify({"status": "ignored"}), 200

        if msg_id in PROCESSADOS:
            return jsonify({"status": "duplicate"}), 200
        PROCESSADOS.add(msg_id)

        if not texto:
            return jsonify({"status": "no_text"}), 200

        # Ignora grupos, newsletters e broadcasts antes de qualquer processamento
        if any(s in from_jid for s in ("@g.us", "@newsletter", "@broadcast")):
            log.info(f"Grupo ignorado: {from_jid}")
            return jsonify({"status": "ignored"}), 200

        # Resolve o @lid (id de privacidade) para o telefone real
        telefone = resolver_telefone(from_jid)

        if len(telefone) > 15 and "@lid" not in from_jid:
            log.info(f"Grupo ignorado: {from_jid}")
            return jsonify({"status": "ignored"}), 200

        log.info(f"Mensagem de {from_jid}: '{texto[:80]}'")

        # Comando administrativo do dono (numero de alertas): "bloquear 5521..."
        if eh_admin(telefone) and processar_comando_admin(from_jid, texto):
            return jsonify({"status": "admin_cmd"}), 200

        if lead_existe_na_planilha(telefone):
            # Se respondeu so o numero do menu (1-5), vira gatilho; senao, classificacao normal
            classificacao = interpretar_opcao_menu(texto) or classificar_resposta(texto)
            atualizar_lead(from_jid, telefone, texto, classificacao)
        else:
            processar_desconhecido(from_jid, telefone, texto)

    except Exception as e:
        log.error(f"Erro ao processar payload: {e}")

    return jsonify({"status": "ok"}), 200


@app.route("/status", methods=["GET"])
def status():
    return jsonify({
        "status": "online",
        "arquivo_leads": ARQUIVO_LEADS,
        "hora": datetime.now().strftime("%d/%m/%Y %H:%M:%S"),
        "processados_sessao": len(PROCESSADOS),
        "conversas_ativas": len(CONVERSAS),
    })


# -- Entrada -------------------------------------------------------------------
if __name__ == "__main__":
    log.info("=" * 60)
    log.info(" Webhook WhatsApp -- Queiroz Seguros")
    log.info(" Status: http://localhost:5000/status")
    log.info("=" * 60)
    app.run(host="0.0.0.0", port=5000, debug=False)
