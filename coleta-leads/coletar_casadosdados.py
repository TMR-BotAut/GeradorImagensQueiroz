#!/usr/bin/env python3
"""
Coletor de leads — Casa dos Dados (CNPJ)
========================================
Percorre TODOS os dias de um ano, aplica os filtros fixos
(Estado=RJ, Cidade=Teresopolis, Situacao=Ativa, Somente MEI, Somente Celular),
usa o MESMO dia em "Data de Abertura - A partir de" e "- Ate", clica em
Pesquisar e coleta de cada resultado: Razao Social, CNPJ e Telefone.

A saida sai no formato da planilha de leads (aba "Leads"), pronta pra colar
no leads.xlsx da campanha:
    Nome = Razao Social | Telefone = 55DDDNUMERO | Cidade = Teresopolis
    Status = pendente   | Observacoes = CNPJ + data de abertura

------------------------------------------------------------------------------
COMO RODAR
    pip install -r requirements.txt
    playwright install chromium
    python coletar_casadosdados.py

RECURSOS
    - Retoma de onde parou (progresso.json) — pode fechar e reabrir.
    - Nao duplica CNPJ ja coletado.
    - Salva a planilha a cada dia (a prova de travamentos).

⚠️ CALIBRAGEM: o site e dinamico e tem protecao anti-robo. Rode com
HEADLESS=False na primeira vez. Se algum campo nao for encontrado, ajuste o
seletor correspondente no bloco SEL abaixo (o README explica como pegar o
seletor certo).
------------------------------------------------------------------------------
"""

from __future__ import annotations

import json
import re
import sys
import time
import random
import unicodedata
from datetime import date, timedelta
from pathlib import Path

from openpyxl import Workbook, load_workbook
from playwright.sync_api import sync_playwright, TimeoutError as PWTimeout

# ════════════════════════════ CONFIG ════════════════════════════
URL = "https://casadosdados.com.br/solucao/cnpj/pesquisa-avancada"

ANO = 2025
DATA_INICIAL = date(ANO, 1, 1)
DATA_FINAL = date(ANO, 12, 31)

ESTADO_DIGITAR = "Rio de Janei"   # o que digitar no campo Estado
ESTADO_OPCAO = "Rio de Janeiro"   # texto da opcao a clicar na lista
CIDADE_DIGITAR = "Teresopolis"    # o que digitar no campo Cidade
CIDADE_OPCAO = "Teres"            # parte do texto da opcao ("Teresopolis - RJ")
CIDADE_SAIDA = "Teresopolis"      # como vai aparecer na planilha

HEADLESS = False                  # False = mostra o navegador (use p/ calibrar)
ARQUIVO_SAIDA = "leads_coletados.xlsx"
ARQUIVO_PROGRESSO = "progresso.json"
ARQUIVO_LOG = "coleta.log"

# Pausas (segundos) — NAO reduza demais; ajudam a nao ser bloqueado
PAUSA_MIN = 1.5
PAUSA_MAX = 3.5
PAUSA_ENTRE_DIAS = 2.0

# Coletar todas as paginas de um dia (True) ou so a primeira (False)
SEGUIR_PAGINACAO = True

# --- SELETORES (AJUSTE AQUI se algo nao for encontrado) -------------
# Dica: rode com HEADLESS=False, abra o DevTools (F12) e use "copiar seletor".
SEL = {
    "estado_input":  "input[placeholder*='Estado'], input[placeholder*='UF']",
    "cidade_input":  "input[placeholder*='Cidade'], input[placeholder*='Munic']",
    "opcao_lista":   ".v-list-item, .autocomplete-item, li[role='option'], .dropdown-item",
    "situacao_ativa": "text=/^\\s*ATIVA\\s*$/i",
    "toggle_mei":     "text=/Somente MEI/i",
    "toggle_celular": "text=/Somente Celular/i",
    "data_de":   "input[placeholder*='partir'], input[placeholder*='Início'], input[name*='inicio']",
    "data_ate":  "input[placeholder*='Até'], input[placeholder*='Ate'], input[name*='fim'], input[name*='final']",
    "btn_pesquisar": "button:has-text('Pesquisar')",
    # Links/itens de cada resultado que levam ao detalhe:
    "detalhe_links": "a[href*='/cnpj/']",
    # Botao de "proxima pagina" (paginacao):
    "btn_proxima": "button:has-text('Próxima'), button[aria-label*='Next'], a[rel='next']",
    "msg_vazio": "text=/nenhum resultado|não encontrad|sem resultados/i",
}
# ═════════════════════════════════════════════════════════════════

COLUNAS = [
    "Nome", "Telefone", "Cidade", "Status", "1º Contato", "Último Contato",
    "Próxima Tentativa", "Tentativas", "Última Resposta", "Hora Resposta",
    "Observações",
]

RE_CNPJ = re.compile(r"\d{2}\.\d{3}\.\d{3}/\d{4}-\d{2}")
RE_TEL = re.compile(r"\(?\d{2}\)?\s?9?\d{4}[-\s]?\d{4}")


# ─────────────────────────── utilitarios ───────────────────────────
def log(msg: str):
    linha = f"{time.strftime('%H:%M:%S')}  {msg}"
    print(linha, flush=True)
    with open(ARQUIVO_LOG, "a", encoding="utf-8") as f:
        f.write(linha + "\n")


def pausa(a: float = PAUSA_MIN, b: float = PAUSA_MAX):
    time.sleep(random.uniform(a, b))


def sem_acento(s: str) -> str:
    return "".join(
        c for c in unicodedata.normalize("NFD", s or "")
        if unicodedata.category(c) != "Mn"
    )


def normalizar_telefone(raw: str) -> str:
    """Converte '(21) 99999-9999' -> '5521999999999'."""
    d = re.sub(r"\D", "", raw or "")
    if not d:
        return ""
    if d.startswith("55") and len(d) >= 12:
        return d
    if len(d) in (10, 11):           # DDD + numero
        return "55" + d
    return d


def extrai_rotulo(texto: str, rotulo: str) -> str:
    """Pega o valor que vem depois de um rotulo (mesma linha ou linha seguinte)."""
    linhas = [l.strip() for l in texto.splitlines()]
    alvo = sem_acento(rotulo).lower()
    for i, l in enumerate(linhas):
        if alvo in sem_acento(l).lower():
            if ":" in l and l.split(":", 1)[1].strip():
                return l.split(":", 1)[1].strip()
            for j in range(i + 1, min(i + 4, len(linhas))):
                if linhas[j]:
                    return linhas[j]
    return ""


# ─────────────────────── planilha de saida ────────────────────────
def carregar_saida():
    """Abre (ou cria) a planilha de saida e devolve (wb, ws, set de cnpjs vistos)."""
    if Path(ARQUIVO_SAIDA).exists():
        wb = load_workbook(ARQUIVO_SAIDA)
        ws = wb["Leads"]
        vistos = set()
        for row in ws.iter_rows(min_row=2, values_only=True):
            obs = str(row[10] or "")
            m = RE_CNPJ.search(obs)
            if m:
                vistos.add(m.group(0))
        log(f"Planilha existente carregada — {len(vistos)} CNPJs ja coletados.")
        return wb, ws, vistos

    wb = Workbook()
    ws = wb.active
    ws.title = "Leads"
    ws.append(COLUNAS)
    wb.save(ARQUIVO_SAIDA)
    return wb, ws, set()


def gravar_lead(ws, razao: str, telefone: str, cnpj: str, dia: date):
    ws.append([
        razao,                       # Nome
        normalizar_telefone(telefone),  # Telefone
        CIDADE_SAIDA,                # Cidade
        "pendente",                  # Status
        None, None, None, 0, None, None,
        f"CNPJ: {cnpj} | Aberto: {dia.strftime('%d/%m/%Y')}",  # Observações
    ])


# ──────────────────────── progresso/retomada ──────────────────────
def carregar_progresso() -> date:
    if Path(ARQUIVO_PROGRESSO).exists():
        try:
            d = json.loads(Path(ARQUIVO_PROGRESSO).read_text())["ultimo_dia"]
            prox = date.fromisoformat(d) + timedelta(days=1)
            log(f"Retomando a partir de {prox.strftime('%d/%m/%Y')}.")
            return max(prox, DATA_INICIAL)
        except Exception:
            pass
    return DATA_INICIAL


def salvar_progresso(dia: date):
    Path(ARQUIVO_PROGRESSO).write_text(
        json.dumps({"ultimo_dia": dia.isoformat()})
    )


# ───────────────────────── acoes no site ──────────────────────────
def selecionar_autocomplete(page, input_sel: str, digitar: str, opcao_txt: str):
    campo = page.locator(input_sel).first
    campo.click()
    campo.fill("")
    campo.type(digitar, delay=120)
    page.wait_for_timeout(1300)
    opcao = page.locator(SEL["opcao_lista"]).filter(has_text=opcao_txt).first
    opcao.click(timeout=6000)
    pausa()


def aplicar_filtros_fixos(page):
    log("Aplicando filtros fixos (RJ, Teresopolis, Ativa, MEI, Celular)...")
    selecionar_autocomplete(page, SEL["estado_input"], ESTADO_DIGITAR, ESTADO_OPCAO)
    selecionar_autocomplete(page, SEL["cidade_input"], CIDADE_DIGITAR, CIDADE_OPCAO)
    for chave in ("situacao_ativa", "toggle_mei", "toggle_celular"):
        try:
            page.locator(SEL[chave]).first.click(timeout=6000)
            pausa(0.6, 1.2)
        except PWTimeout:
            log(f"  ⚠ Nao achei '{chave}' — confira o seletor no bloco SEL.")
    log("Filtros fixos aplicados.")


def definir_datas(page, dia: date):
    valor = dia.strftime("%d/%m/%Y")
    for chave in ("data_de", "data_ate"):
        campo = page.locator(SEL[chave]).first
        campo.click()
        campo.fill("")
        campo.type(valor, delay=60)
        page.keyboard.press("Escape")  # fecha eventual calendario
        pausa(0.4, 0.9)


def coletar_detalhes(context, page) -> list[dict]:
    """Coleta os detalhes de cada resultado da pagina atual."""
    page.wait_for_timeout(1500)

    if page.locator(SEL["msg_vazio"]).count() > 0:
        return []

    # Pega os links de detalhe (unicos) da area de resultados
    hrefs = []
    for el in page.locator(SEL["detalhe_links"]).all():
        h = el.get_attribute("href")
        if h and "/cnpj/" in h and h not in hrefs:
            hrefs.append(h)

    if not hrefs:
        return []

    base = "https://casadosdados.com.br"
    coletados = []
    for href in hrefs:
        url = href if href.startswith("http") else base + href
        det = context.new_page()
        try:
            det.goto(url, wait_until="domcontentloaded", timeout=30000)
            det.wait_for_timeout(900)
            texto = det.inner_text("body")
            cnpj_m = RE_CNPJ.search(texto)
            cnpj = cnpj_m.group(0) if cnpj_m else ""
            razao = extrai_rotulo(texto, "Razão Social") or extrai_rotulo(texto, "Nome")
            tel = extrai_rotulo(texto, "Telefone")
            if not RE_TEL.search(tel or ""):
                m = RE_TEL.search(texto)
                tel = m.group(0) if m else ""
            if cnpj:
                coletados.append({"cnpj": cnpj, "razao": razao.strip(), "telefone": tel.strip()})
        except PWTimeout:
            log(f"  ⚠ timeout no detalhe: {url}")
        finally:
            det.close()
            pausa()
    return coletados


# ──────────────────────────── principal ───────────────────────────
def main():
    wb, ws, vistos = carregar_saida()
    inicio = carregar_progresso()

    with sync_playwright() as p:
        navegador = p.chromium.launch(headless=HEADLESS)
        context = navegador.new_context(
            locale="pt-BR",
            user_agent=("Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                        "AppleWebKit/537.36 (KHTML, like Gecko) "
                        "Chrome/122.0 Safari/537.36"),
        )
        page = context.new_page()
        log(f"Abrindo {URL}")
        page.goto(URL, wait_until="domcontentloaded", timeout=60000)

        if not HEADLESS:
            input(">>> Se aparecer captcha/aviso de cookies, resolva no navegador "
                  "e tecle ENTER aqui para começar...")

        aplicar_filtros_fixos(page)

        dia = inicio
        total_novos = 0
        while dia <= DATA_FINAL:
            definir_datas(page, dia)
            page.locator(SEL["btn_pesquisar"]).first.click()
            log(f"🔎 {dia.strftime('%d/%m/%Y')} — pesquisando...")
            try:
                page.wait_for_load_state("networkidle", timeout=15000)
            except PWTimeout:
                pass

            novos_dia = 0
            while True:
                for d in coletar_detalhes(context, page):
                    if d["cnpj"] in vistos:
                        continue
                    vistos.add(d["cnpj"])
                    gravar_lead(ws, d["razao"], d["telefone"], d["cnpj"], dia)
                    novos_dia += 1
                    total_novos += 1

                if not SEGUIR_PAGINACAO:
                    break
                prox = page.locator(SEL["btn_proxima"]).first
                if prox.count() == 0 or not prox.is_enabled():
                    break
                prox.click()
                page.wait_for_timeout(1500)

            wb.save(ARQUIVO_SAIDA)
            salvar_progresso(dia)
            log(f"   {dia.strftime('%d/%m/%Y')}: +{novos_dia} novos (total {total_novos})")

            dia += timedelta(days=1)
            time.sleep(PAUSA_ENTRE_DIAS)

        navegador.close()

    log(f"✅ Concluido. {total_novos} leads novos em '{ARQUIVO_SAIDA}'.")
    log("   Copie as linhas (a partir da linha 2) para a aba 'Leads' do seu leads.xlsx.")


if __name__ == "__main__":
    try:
        main()
    except KeyboardInterrupt:
        print("\nInterrompido. Rode de novo para retomar de onde parou.")
        sys.exit(0)
